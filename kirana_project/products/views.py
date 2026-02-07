from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.http import HttpResponse
from django.db.models import Sum, F
from .models import Product, ProductCategory
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime
from decimal import Decimal

# Create your views here.
def product_list(request):
    products = Product.objects.all().order_by('name')
    total_products = products.count()
    low_stock_items = sum(1 for p in products if p.stock <= p.min_stock_level)
    total_categories = ProductCategory.objects.count()
    total_value = products.aggregate(total=Sum(F('stock') * F('price')))['total'] or 0
    return render(request, 'products/product_list.html', {
        'products': products,
        'total_products': total_products,
        'low_stock_items': low_stock_items,
        'total_categories': total_categories,
        'total_value': total_value,
    })

def add_product(request):
    if request.method == 'POST':
        try:
            # Get form data
            name = request.POST.get('name')
            brand = request.POST.get('brand')
            barcode = request.POST.get('barcode')
            price = request.POST.get('price')
            cost_price = request.POST.get('cost_price')
            gst_rate = request.POST.get('gst_rate') or 0
            stock = request.POST.get('stock') or 0
            min_stock_level = request.POST.get('min_stock_level') or 5
            unit = request.POST.get('unit') or 'pcs'
            category_id = request.POST.get('category')
            category = ProductCategory.objects.get(pk=category_id) if category_id else None

            # Create and save product
            product = Product.objects.create(
                name=name,
                brand=brand,
                barcode=barcode,
                price=float(price),
                cost_price=float(cost_price) if cost_price else 0,
                gst_rate=float(gst_rate),
                stock=int(stock),
                min_stock_level=int(min_stock_level),
                unit=unit,
                category=category
            )
            
            messages.success(request, f'Product "{name}" added successfully!')
            return redirect('products:product_list')
        except Exception as e:
            messages.error(request, f'Error adding product: {str(e)}')
    categories = ProductCategory.objects.all()
    if not categories.exists():
        default_categories = [
            "Groceries", "Beverages", "Snacks", "Personal Care", "Household"
        ]
        for cat_name in default_categories:
            ProductCategory.objects.create(name=cat_name)
        categories = ProductCategory.objects.all()
    return render(request, 'products/product_form.html', {'categories': categories})

def product_detail(request, pk):
    product = get_object_or_404(Product, pk=pk)
    return render(request, 'products/product_detail.html', {'product': product})

def edit_product(request, pk):
    product = get_object_or_404(Product, pk=pk)
    categories = ProductCategory.objects.all()
    if not categories.exists():
        default_categories = [
            "Groceries", "Beverages", "Snacks", "Personal Care", "Household"
        ]
        for cat_name in default_categories:
            ProductCategory.objects.create(name=cat_name)
        categories = ProductCategory.objects.all()
    if request.method == 'POST':
        try:
            # Get form data
            product.name = request.POST.get('name')
            product.brand = request.POST.get('brand')
            product.barcode = request.POST.get('barcode')
            product.price = float(request.POST.get('price'))
            product.cost_price = float(request.POST.get('cost_price')) if request.POST.get('cost_price') else 0
            product.gst_rate = float(request.POST.get('gst_rate')) if request.POST.get('gst_rate') else 0
            product.stock = int(request.POST.get('stock')) if request.POST.get('stock') else 0
            product.min_stock_level = int(request.POST.get('min_stock_level')) if request.POST.get('min_stock_level') else 5
            product.unit = request.POST.get('unit') or 'pcs'
            category_id = request.POST.get('category')
            product.category = ProductCategory.objects.get(pk=category_id) if category_id else None
            product.save()
            messages.success(request, f'Product "{product.name}" updated successfully!')
            return redirect('products:product_detail', pk=pk)
        except Exception as e:
            messages.error(request, f'Error updating product: {str(e)}')
    return render(request, 'products/product_form.html', {'product': product, 'categories': categories})

def delete_product(request, pk):
    product = get_object_or_404(Product, pk=pk)
    if request.method == 'POST':
        try:
            product_name = product.name
            product.delete()
            messages.success(request, f'Product "{product_name}" deleted successfully!')
            return redirect('products:product_list')
        except Exception as e:
            messages.error(request, f'Error deleting product: {str(e)}')
            return redirect('products:product_detail', pk=pk)
    return render(request, 'products/product_confirm_delete.html', {'product': product})

def stock_report(request):
    from .models import Product
    from django.db.models import F
    
    # Get fresh product data from database
    products = Product.objects.all().order_by('name')
    
    # Calculate stock statistics
    out_of_stock = products.filter(stock=0).count()
    low_stock = products.filter(stock__gt=0, stock__lte=F('min_stock_level')).count()
    good_stock = products.filter(stock__gt=F('min_stock_level')).count()
    
    # Calculate inventory value
    total_inventory_value = 0
    for product in products:
        product.stock_value = product.stock * product.cost_price
        total_inventory_value += product.stock_value
    
    context = {
        'out_of_stock': out_of_stock,
        'low_stock': low_stock,
        'good_stock': good_stock,
        'products': products,
        'total_inventory_value': total_inventory_value,
    }
    return render(request, 'products/stock_report.html', context)

def category_list(request):
    return render(request, 'products/category_list.html')

def export_stock_report(request):
    """Export stock report data to Excel file"""
    
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Stock Report"
    
    # Set up headers
    headers = [
        'Product Name', 'Brand', 'Barcode', 'Category', 'Unit', 'Current Stock', 
        'Min Stock Level', 'Cost Price', 'Selling Price', 'Stock Value', 'GST Rate', 'Stock Status'
    ]
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Get all products data
    products = Product.objects.all().order_by('name')
    
    # Calculate statistics
    out_of_stock = products.filter(stock=0).count()
    low_stock = products.filter(stock__gt=0, stock__lte=F('min_stock_level')).count()
    good_stock = products.filter(stock__gt=F('min_stock_level')).count()
    total_inventory_value = Decimal('0')
    
    # Write product data
    row = 2
    for product in products:
        # Calculate stock value
        stock_value = product.stock * product.cost_price
        total_inventory_value += stock_value
        
        # Determine stock status
        if product.stock == 0:
            stock_status = "Out of Stock"
        elif product.stock <= product.min_stock_level:
            stock_status = "Low Stock"
        else:
            stock_status = "Good Stock"
        
        # Write row data
        ws.cell(row=row, column=1, value=product.name)
        ws.cell(row=row, column=2, value=product.brand or "")
        ws.cell(row=row, column=3, value=product.barcode or "")
        ws.cell(row=row, column=4, value=product.category.name if product.category else "")
        ws.cell(row=row, column=5, value=product.unit)
        ws.cell(row=row, column=6, value=product.stock)
        ws.cell(row=row, column=7, value=product.min_stock_level)
        ws.cell(row=row, column=8, value=float(product.cost_price))
        ws.cell(row=row, column=9, value=float(product.price))
        ws.cell(row=row, column=10, value=float(stock_value))
        ws.cell(row=row, column=11, value=f"{product.gst_rate}%")
        ws.cell(row=row, column=12, value=stock_status)
        
        row += 1
    
    # Adjust column widths
    column_widths = [25, 15, 15, 15, 8, 12, 15, 12, 12, 12, 10, 15]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Add a summary row
    summary_row = row + 2
    ws.cell(row=summary_row, column=1, value="SUMMARY").font = Font(bold=True)
    ws.cell(row=summary_row + 1, column=1, value=f"Total Products: {products.count()}")
    ws.cell(row=summary_row + 2, column=1, value=f"Out of Stock: {out_of_stock}")
    ws.cell(row=summary_row + 3, column=1, value=f"Low Stock: {low_stock}")
    ws.cell(row=summary_row + 4, column=1, value=f"Good Stock: {good_stock}")
    ws.cell(row=summary_row + 5, column=1, value=f"Total Inventory Value: ₹{total_inventory_value:.2f}")
    ws.cell(row=summary_row + 6, column=1, value=f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    
    # Create HTTP response with Excel file
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    # Set filename with current date
    filename = f"stock_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Save workbook to response
    wb.save(response)
    
    return response
