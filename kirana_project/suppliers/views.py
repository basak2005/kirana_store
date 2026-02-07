from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.db import transaction, models
from django.db.models import F
from django.views.decorators.http import require_POST
from django.http import HttpResponse
from .models import Supplier, PurchaseOrder, PurchaseItem, Expense
from products.models import Product
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime

# Create your views here.

def delete_purchase_order(request, pk):
    order = get_object_or_404(PurchaseOrder, pk=pk)
    if request.method == 'POST':
        try:
            order.delete()
            messages.success(request, f'Purchase Order #{order.po_number} deleted successfully!')
            return redirect('suppliers:purchase_orders')
        except Exception as e:
            messages.error(request, f'Error deleting purchase order: {str(e)}')
            return redirect('suppliers:purchase_detail', pk=pk)
    return render(request, 'suppliers/purchase_detail.html', {'order': order})
from django.views.decorators.http import require_POST
@require_POST
def mark_as_completed(request, pk):
    order = get_object_or_404(PurchaseOrder, pk=pk)
    
    if order.status == 'completed':
        messages.info(request, f'Purchase Order #{order.po_number} is already completed.')
        return redirect('suppliers:purchase_detail', pk=order.pk)
    
    try:
        with transaction.atomic():
            updated_products = []
            
            # Update stock for each item in the purchase order
            for item in order.items.all():
                product = item.product
                old_stock = product.stock
                
                # Update stock using F() expression for atomic update
                Product.objects.filter(pk=product.pk).update(stock=F('stock') + item.quantity)
                
                # Refresh the product from database to get updated stock
                product.refresh_from_db()
                
                updated_products.append({
                    'name': product.name,
                    'old_stock': old_stock,
                    'quantity_added': item.quantity,
                    'new_stock': product.stock
                })
            
            # Mark the order as completed
            order.status = 'completed'
            order.save(update_fields=['status'])
            
            # Create a detailed success message
            product_updates = ", ".join([f"{p['name']}: {p['old_stock']} → {p['new_stock']}" for p in updated_products])
            messages.success(request, f'Purchase Order #{order.po_number} marked as completed. Stock updated: {product_updates}')
            
    except Exception as e:
        messages.error(request, f'Error updating stock: {str(e)}')
    
    return redirect('suppliers:purchase_detail', pk=order.pk)

def supplier_list(request):
    suppliers = Supplier.objects.all().order_by('name')
    total_suppliers = suppliers.count()
    active_suppliers = suppliers.filter(is_active=True).count()
    inactive_suppliers = suppliers.filter(is_active=False).count()
    # For pending orders and total purchases, you can add logic if needed
    pending_orders = PurchaseOrder.objects.filter(status='pending').count() if hasattr(PurchaseOrder, 'status') else 0
    total_purchases = PurchaseOrder.objects.aggregate(total=models.Sum('total_amount'))['total'] or 0
    return render(request, 'suppliers/supplier_list.html', {
        'suppliers': suppliers,
        'total_suppliers': total_suppliers,
        'active_suppliers': active_suppliers,
        'inactive_suppliers': inactive_suppliers,
        'pending_orders': pending_orders,
        'total_purchases': total_purchases,
    })

def add_supplier(request):
    if request.method == 'POST':
        try:
            # Get form data
            name = request.POST.get('name')
            contact_person = request.POST.get('contact_person', '')
            phone = request.POST.get('phone')
            email = request.POST.get('email', '')
            address = request.POST.get('address')
            gst_number = request.POST.get('gst_number', '')
            is_active = request.POST.get('is_active') == 'on'
            
            # Validate required fields
            if not name or not phone or not address:
                messages.error(request, 'Please fill in all required fields (Name, Phone, Address).')
                return render(request, 'suppliers/supplier_form.html')
            
            # Create and save supplier
            supplier = Supplier.objects.create(
                name=name,
                contact_person=contact_person,
                phone=phone,
                email=email if email else None,
                address=address,
                gst_number=gst_number,
                is_active=is_active
            )
            
            messages.success(request, f'Supplier "{name}" added successfully!')
            return redirect('suppliers:supplier_list')
        except Exception as e:
            messages.error(request, f'Error adding supplier: {str(e)}')
    return render(request, 'suppliers/supplier_form.html')

def supplier_detail(request, pk):
    supplier = get_object_or_404(Supplier, pk=pk)
    return render(request, 'suppliers/supplier_detail.html', {'supplier': supplier})

def edit_supplier(request, pk):
    supplier = get_object_or_404(Supplier, pk=pk)
    if request.method == 'POST':
        try:
            # Get form data
            supplier.name = request.POST.get('name')
            supplier.contact_person = request.POST.get('contact_person', '')
            supplier.phone = request.POST.get('phone')
            supplier.email = request.POST.get('email', '') or None
            supplier.address = request.POST.get('address')
            supplier.gst_number = request.POST.get('gst_number', '')
            supplier.is_active = request.POST.get('is_active') == 'on'
            
            # Validate required fields
            if not supplier.name or not supplier.phone or not supplier.address:
                messages.error(request, 'Please fill in all required fields (Name, Phone, Address).')
                return render(request, 'suppliers/supplier_form.html', {'supplier': supplier})
            
            supplier.save()
            
            messages.success(request, f'Supplier "{supplier.name}" updated successfully!')
            return redirect('suppliers:supplier_detail', pk=pk)
        except Exception as e:
            messages.error(request, f'Error updating supplier: {str(e)}')
    return render(request, 'suppliers/supplier_form.html', {'supplier': supplier})

def delete_supplier(request, pk):
    supplier = get_object_or_404(Supplier, pk=pk)
    if request.method == 'POST':
        try:
            supplier_name = supplier.name
            supplier.delete()
            messages.success(request, f'Supplier "{supplier_name}" deleted successfully!')
            return redirect('suppliers:supplier_list')
        except Exception as e:
            messages.error(request, f'Error deleting supplier: {str(e)}')
            return redirect('suppliers:supplier_detail', pk=pk)
    return render(request, 'suppliers/supplier_confirm_delete.html', {'supplier': supplier})

def purchase_orders(request):
    from django.db.models import Sum, Count, Q
    from datetime import datetime, date
    from django.utils import timezone
    
    try:
        # Get all orders with related data and annotate with total items quantity
        all_orders = PurchaseOrder.objects.select_related('supplier').prefetch_related('items').annotate(
            total_items_quantity=Sum('items__quantity')
        ).all()
        
        # Calculate OVERALL statistics (not affected by filters)
        total_orders = all_orders.count()
        pending_orders = all_orders.filter(status='pending').count()
        completed_orders = all_orders.filter(status='completed').count()
        cancelled_orders = all_orders.filter(status='cancelled').count()
        
        # Calculate totals for ALL orders
        total_value = all_orders.aggregate(total=Sum('total_amount'))['total'] or 0
        
        # This month's orders (all orders, not filtered)
        current_month = timezone.now().month
        current_year = timezone.now().year
        monthly_value = all_orders.filter(
            date__year=current_year,
            date__month=current_month
        ).aggregate(total=Sum('total_amount'))['total'] or 0
        
        # Calculate total items across all orders using efficient query
        from suppliers.models import PurchaseItem
        total_items = PurchaseItem.objects.filter(
            order__in=all_orders
        ).aggregate(total=Sum('quantity'))['total'] or 0
        
        # Get suppliers for filter dropdown
        suppliers = Supplier.objects.filter(is_active=True).order_by('name')
        
        # Now apply filters to create the filtered orders for display
        filtered_orders = all_orders.order_by('-date')
        
        # Apply filters if provided
        status_filter = request.GET.get('status')
        supplier_filter = request.GET.get('supplier')
        date_filter = request.GET.get('date')
        
        if status_filter and status_filter != 'all':
            filtered_orders = filtered_orders.filter(status=status_filter)
        
        if supplier_filter and supplier_filter != 'all':
            try:
                filtered_orders = filtered_orders.filter(supplier_id=int(supplier_filter))
            except (ValueError, TypeError):
                pass  # Invalid supplier ID, ignore filter
        
        if date_filter:
            try:
                filter_date = datetime.strptime(date_filter, '%Y-%m-%d').date()
                filtered_orders = filtered_orders.filter(date__date=filter_date)
            except ValueError:
                pass  # Invalid date format, ignore filter
        
        context = {
            'orders': filtered_orders,
            'total_orders': total_orders,
            'pending_orders': pending_orders,
            'completed_orders': completed_orders,
            'cancelled_orders': cancelled_orders,
            'total_value': total_value,
            'monthly_value': monthly_value,
            'total_items': total_items,
            'suppliers': suppliers,
            'today': date.today(),
            'status_filter': status_filter,
            'supplier_filter': supplier_filter,
            'date_filter': date_filter,
        }
        
        return render(request, 'suppliers/purchase_orders.html', context)
        
    except Exception as e:
        messages.error(request, f'Error loading purchase orders: {str(e)}')
        print(f"Error in purchase_orders view: {str(e)}")
        return render(request, 'suppliers/purchase_orders.html', {
            'orders': [],
            'total_orders': 0,
            'pending_orders': 0,
            'completed_orders': 0,
            'cancelled_orders': 0,
            'total_value': 0,
            'monthly_value': 0,
            'total_items': 0,
            'suppliers': [],
            'today': date.today(),
        })

def new_purchase(request):
    if request.method == 'POST':
        try:
            supplier_id = request.POST.get('supplier')
            supplier = Supplier.objects.get(pk=supplier_id)
            expected_delivery = request.POST.get('expected_delivery') or None
            notes = request.POST.get('notes', '')

            order = PurchaseOrder.objects.create(
                supplier=supplier,
                total_amount=0,
                subtotal=0,
                total_gst=0,
                expected_delivery=expected_delivery,
                notes=notes,
                status='pending'
            )

            products_ids = request.POST.getlist('product[]')
            quantities = request.POST.getlist('quantity[]')
            costs = request.POST.getlist('cost[]')
            gsts = request.POST.getlist('gst[]')

            subtotal = 0
            total_gst = 0
            total_amount = 0

            from products.models import Product
            for idx, prod_id in enumerate(products_ids):
                if not prod_id:
                    continue
                product = Product.objects.get(pk=prod_id)
                quantity = int(quantities[idx]) if idx < len(quantities) else 1
                unit_cost = float(costs[idx]) if idx < len(costs) and costs[idx] else float(product.cost_price)
                gst_rate = float(gsts[idx]) if idx < len(gsts) and gsts[idx] else float(product.gst_rate)

                base_amount = quantity * unit_cost
                gst_amount = (base_amount * gst_rate) / 100
                total_cost = base_amount + gst_amount

                PurchaseItem.objects.create(
                    order=order,
                    product=product,
                    quantity=quantity,
                    unit_cost=unit_cost,
                    gst_rate=gst_rate,
                    gst_amount=gst_amount,
                    total_cost=total_cost
                )

                # Stock will be updated when PO is marked as completed
                # Do not update stock here to avoid double updates

                subtotal += base_amount
                total_gst += gst_amount
                total_amount += total_cost

            order.subtotal = subtotal
            order.total_gst = total_gst
            order.total_amount = total_amount
            order.save()

            messages.success(request, f'Purchase Order #{order.po_number} created successfully!')
            return redirect('suppliers:purchase_detail', pk=order.pk)
        except Exception as e:
            messages.error(request, f'Error creating purchase order: {str(e)}')

    from products.models import Product
    suppliers = Supplier.objects.all()
    products = Product.objects.filter(is_active=True)
    return render(request, 'suppliers/purchase_form.html', {'suppliers': suppliers, 'products': products})

def edit_purchase(request, pk):
    order = get_object_or_404(PurchaseOrder, pk=pk)
    
    # Only allow editing of pending orders
    if order.status != 'pending':
        messages.error(request, 'Only pending purchase orders can be edited.')
        return redirect('suppliers:purchase_detail', pk=pk)
    
    if request.method == 'POST':
        try:
            supplier_id = request.POST.get('supplier')
            supplier = Supplier.objects.get(pk=supplier_id)
            expected_delivery = request.POST.get('expected_delivery') or None
            notes = request.POST.get('notes', '')

            # Update order basic info
            order.supplier = supplier
            order.expected_delivery = expected_delivery
            order.notes = notes

            # Delete existing items and recreate them
            order.items.all().delete()

            products_ids = request.POST.getlist('product[]')
            quantities = request.POST.getlist('quantity[]')
            costs = request.POST.getlist('cost[]')
            gsts = request.POST.getlist('gst[]')

            subtotal = 0
            total_gst = 0
            total_amount = 0

            from products.models import Product
            for idx, prod_id in enumerate(products_ids):
                if not prod_id:
                    continue
                product = Product.objects.get(pk=prod_id)
                quantity = int(quantities[idx]) if idx < len(quantities) else 1
                unit_cost = float(costs[idx]) if idx < len(costs) and costs[idx] else float(product.cost_price)
                gst_rate = float(gsts[idx]) if idx < len(gsts) and gsts[idx] else float(product.gst_rate)

                base_amount = quantity * unit_cost
                gst_amount = (base_amount * gst_rate) / 100
                total_cost = base_amount + gst_amount

                PurchaseItem.objects.create(
                    order=order,
                    product=product,
                    quantity=quantity,
                    unit_cost=unit_cost,
                    gst_rate=gst_rate,
                    gst_amount=gst_amount,
                    total_cost=total_cost
                )

                subtotal += base_amount
                total_gst += gst_amount
                total_amount += total_cost

            order.subtotal = subtotal
            order.total_gst = total_gst
            order.total_amount = total_amount
            order.save()

            messages.success(request, f'Purchase Order #{order.po_number} updated successfully!')
            return redirect('suppliers:purchase_detail', pk=order.pk)
        except Exception as e:
            messages.error(request, f'Error updating purchase order: {str(e)}')

    from products.models import Product
    suppliers = Supplier.objects.all()
    products = Product.objects.filter(is_active=True)
    context = {
        'suppliers': suppliers, 
        'products': products, 
        'order': order,
        'is_edit': True
    }
    return render(request, 'suppliers/purchase_form.html', context)

def purchase_detail(request, pk):
    order = get_object_or_404(PurchaseOrder, pk=pk)
    return render(request, 'suppliers/purchase_detail.html', {'order': order})

def expense_list(request):
    expenses = Expense.objects.all().order_by('-created_at')
    return render(request, 'suppliers/expense_list.html', {'expenses': expenses})

def add_expense(request):
    return render(request, 'suppliers/expense_form.html')

def export_purchase_orders(request):
    """Export purchase orders data to Excel file"""
    
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Purchase Orders Report"
    
    # Set up headers
    headers = [
        'PO Number', 'Date', 'Supplier', 'Item Names', 'Total Items', 
        'Subtotal', 'Total GST', 'Total Amount', 'Status', 'Expected Delivery', 'Notes'
    ]
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Get all purchase orders data
    orders = PurchaseOrder.objects.all().order_by('-date')
    
    # Write purchase orders data
    row = 2
    for order in orders:
        # Get items for this order
        items_text = ", ".join([f"{item.product.name} (x{item.quantity})" for item in order.items.all()])
        
        # Calculate total items quantity
        total_items = sum(item.quantity for item in order.items.all())
        
        # Write row data
        ws.cell(row=row, column=1, value=order.po_number)
        ws.cell(row=row, column=2, value=order.date.strftime('%Y-%m-%d'))
        ws.cell(row=row, column=3, value=order.supplier.name)
        ws.cell(row=row, column=4, value=items_text)
        ws.cell(row=row, column=5, value=total_items)
        ws.cell(row=row, column=6, value=float(order.subtotal))
        ws.cell(row=row, column=7, value=float(order.total_gst))
        ws.cell(row=row, column=8, value=float(order.total_amount))
        ws.cell(row=row, column=9, value=order.get_status_display())
        ws.cell(row=row, column=10, value=order.expected_delivery.strftime('%Y-%m-%d') if order.expected_delivery else "")
        ws.cell(row=row, column=11, value=order.notes or "")
        
        row += 1
    
    # Adjust column widths
    column_widths = [15, 12, 20, 40, 12, 12, 12, 12, 12, 15, 30]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Add a summary row
    summary_row = row + 2
    ws.cell(row=summary_row, column=1, value="SUMMARY").font = Font(bold=True)
    ws.cell(row=summary_row + 1, column=1, value=f"Total Purchase Orders: {orders.count()}")
    ws.cell(row=summary_row + 2, column=1, value=f"Total Amount: ₹{sum(order.total_amount for order in orders):.2f}")
    ws.cell(row=summary_row + 3, column=1, value=f"Pending Orders: {orders.filter(status='pending').count()}")
    ws.cell(row=summary_row + 4, column=1, value=f"Completed Orders: {orders.filter(status='completed').count()}")
    ws.cell(row=summary_row + 5, column=1, value=f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    
    # Create HTTP response with Excel file
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    # Set filename with current date
    filename = f"purchase_orders_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Save workbook to response
    wb.save(response)
    
    return response


def print_purchase_order(request, pk):
    order = get_object_or_404(PurchaseOrder, pk=pk)
    return render(request, 'suppliers/purchase_invoice.html', {'order': order})
