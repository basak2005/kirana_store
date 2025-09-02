from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.db import transaction
from django.db.models import F
from .models import Sale, SaleItem
from customers.models import Customer
from products.models import Product
from decimal import Decimal
import json
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime

# Create your views here.
def sale_list(request):
    sales = Sale.objects.all().order_by('-date')
    return render(request, 'sales/sale_list.html', {'sales': sales})

def new_sale(request):
    if request.method == 'POST':
        try:
            with transaction.atomic():  # Ensure all operations succeed or fail together
                # Get customer
                customer_id = request.POST.get('customer')
                customer = None
                if customer_id:
                    customer = Customer.objects.get(pk=customer_id)
                
                payment_mode = request.POST.get('payment_mode', 'cash')
                is_credit = payment_mode == 'credit'
                
                # Create sale
                sale = Sale.objects.create(
                    customer=customer,
                    payment_mode=payment_mode,
                    is_credit=is_credit,
                    total_amount=0  # Will be calculated below
                )
                
                # Process items
                products = request.POST.getlist('product[]')
                quantities = request.POST.getlist('quantity[]')
                prices = request.POST.getlist('price[]')
                gsts = request.POST.getlist('gst[]')
                
                total_amount = Decimal('0.00')
                subtotal = Decimal('0.00')
                total_gst = Decimal('0.00')
                
                updated_products = []  # Track stock updates for logging
                
                for i, product_id in enumerate(products):
                    if product_id and i < len(quantities) and i < len(prices):
                        try:
                            product = Product.objects.select_for_update().get(pk=product_id)  # Lock for update
                            quantity = int(quantities[i])
                            unit_price = Decimal(prices[i]) if prices[i] else Decimal(str(product.price))
                            gst_rate = Decimal(gsts[i]) if i < len(gsts) and gsts[i] else Decimal(str(product.gst_rate))
                            
                            # Stock check
                            if product.stock < quantity:
                                messages.error(request, f"{product.name} has insufficient stock! Available: {product.stock}, Required: {quantity}")
                                return redirect('sales:new_sale')
                            
                            # Calculate totals
                            line_total = quantity * unit_price
                            gst_amount = (line_total * gst_rate) / Decimal('100')
                            total_price = line_total + gst_amount
                            
                            # Create sale item
                            SaleItem.objects.create(
                                sale=sale,
                                product=product,
                                quantity=quantity,
                                unit_price=unit_price,
                                gst_rate=gst_rate,
                                gst_amount=gst_amount,
                                total_price=total_price
                            )
                            
                            # Update product stock atomically
                            old_stock = product.stock
                            Product.objects.filter(pk=product.pk).update(stock=F('stock') - quantity)
                            
                            # Refresh to get updated stock for logging
                            product.refresh_from_db()
                            updated_products.append({
                                'name': product.name,
                                'old_stock': old_stock,
                                'quantity_sold': quantity,
                                'new_stock': product.stock
                            })
                            
                            subtotal += line_total
                            total_gst += gst_amount
                            total_amount += total_price
                        except (Product.DoesNotExist, ValueError) as e:
                            messages.error(request, f'Error processing product: {str(e)}')
                            return redirect('sales:new_sale')
                
                # Update sale totals
                sale.subtotal = subtotal
                sale.total_gst = total_gst
                sale.total_amount = total_amount
                sale.save()
                
                # Update customer credit if sale is credit
                if sale.is_credit and sale.customer:
                    from customers.models import CustomerCredit
                    credit_obj, created = CustomerCredit.objects.get_or_create(customer=sale.customer)
                    credit_obj.total_credit += sale.total_amount
                    credit_obj.save()
                
                # Update daily report
                from reports.models import DailyReport
                from datetime import date
                try:
                    DailyReport.generate_report(date.today())
                except Exception as report_error:
                    print(f"Warning: Could not update daily report: {report_error}")
                
                # Log stock updates
                stock_updates = ", ".join([f"{p['name']}: {p['old_stock']} → {p['new_stock']}" for p in updated_products])
                print(f"Sale #{sale.invoice_number} completed. Stock updated: {stock_updates}")
                
                messages.success(request, f'Sale #{sale.invoice_number} created successfully! Stock updated for {len(updated_products)} products.')
                return redirect('sales:sale_detail', pk=sale.pk)
                
        except Exception as e:
            messages.error(request, f'Error creating sale: {str(e)}')
            print(f"Error in new_sale: {str(e)}")
    
    customers = Customer.objects.all()
    products = Product.objects.filter(is_active=True, stock__gt=0).order_by('name')  # Only show products with stock
    return render(request, 'sales/sale_form.html', {
        'customers': customers,
        'products': products
    })

def sale_detail(request, pk):
    sale = get_object_or_404(Sale, pk=pk)
    # Calculate total profit for this sale
    total_profit = sum([item.profit for item in sale.items.all()])
    return render(request, 'sales/sale_detail.html', {'sale': sale, 'total_profit': total_profit})

def print_invoice(request, pk):
    sale = get_object_or_404(Sale, pk=pk)
    return render(request, 'sales/invoice.html', {'sale': sale})

def edit_sale(request, pk):
    sale = get_object_or_404(Sale, pk=pk)
    customers = Customer.objects.all()
    products = Product.objects.all()
    return render(request, 'sales/sale_form.html', {
        'sale': sale,
        'customers': customers,
        'products': products
    })

def delete_sale(request, pk):
    sale = get_object_or_404(Sale, pk=pk)
    if request.method == 'POST':
        try:
            with transaction.atomic():
                sale_number = sale.invoice_number
                restored_products = []
                
                # Restore stock for all items in the sale
                for item in sale.items.all():
                    product = item.product
                    old_stock = product.stock
                    
                    # Restore stock atomically
                    Product.objects.filter(pk=product.pk).update(stock=F('stock') + item.quantity)
                    
                    # Refresh to get updated stock for logging
                    product.refresh_from_db()
                    restored_products.append({
                        'name': product.name,
                        'old_stock': old_stock,
                        'quantity_restored': item.quantity,
                        'new_stock': product.stock
                    })
                
                # Restore customer credit if it was a credit sale
                if sale.is_credit and sale.customer:
                    from customers.models import CustomerCredit
                    try:
                        credit_obj = CustomerCredit.objects.get(customer=sale.customer)
                        credit_obj.total_credit -= sale.total_amount
                        if credit_obj.total_credit < 0:
                            credit_obj.total_credit = 0  # Don't allow negative credit
                        credit_obj.save()
                    except CustomerCredit.DoesNotExist:
                        pass  # Credit record doesn't exist, nothing to restore
                
                # Delete the sale
                sale.delete()
                
                # Update daily report
                from reports.models import DailyReport
                from datetime import date
                try:
                    DailyReport.generate_report(date.today())
                except Exception as report_error:
                    print(f"Warning: Could not update daily report: {report_error}")
                
                # Log stock restoration
                stock_restorations = ", ".join([f"{p['name']}: {p['old_stock']} → {p['new_stock']}" for p in restored_products])
                print(f"Sale #{sale_number} deleted. Stock restored: {stock_restorations}")
                
                messages.success(request, f'Sale #{sale_number} deleted successfully! Stock restored for {len(restored_products)} products.')
                return redirect('sales:sale_list')
        except Exception as e:
            messages.error(request, f'Error deleting sale: {str(e)}')
            print(f"Error in delete_sale: {str(e)}")
            return redirect('sales:sale_detail', pk=pk)
    return render(request, 'sales/sale_confirm_delete.html', {'sale': sale})

def today_sales(request):
    return render(request, 'sales/today_sales.html')

def mark_credit_paid(request, pk):
    sale = get_object_or_404(Sale, pk=pk)
    if sale.is_credit and not sale.credit_paid:
        sale.credit_paid = True
        sale.credit_paid_date = timezone.now().date()
        sale.save()
        messages.success(request, "Credit payment recorded.")
    return redirect('sales:sale_detail', pk=pk)

def export_sales(request):
    """Export sales data to Excel file"""
    
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Report"
    
    # Set up headers
    headers = [
        'Invoice #', 'Date', 'Time', 'Customer', 'Payment Mode', 
        'Credit Status', 'Items', 'Total Amount', 'Notes'
    ]
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Get all sales data
    sales = Sale.objects.all().order_by('-date')
    
    # Write sales data
    row = 2
    for sale in sales:
        # Get items for this sale
        items_text = ", ".join([f"{item.product.name} (x{item.quantity})" for item in sale.items.all()])
        
        # Determine credit status
        credit_status = ""
        if sale.is_credit:
            credit_status = "Paid" if sale.credit_paid else "Unpaid"
        else:
            credit_status = "N/A"
        
        # Write row data
        ws.cell(row=row, column=1, value=sale.invoice_number)
        ws.cell(row=row, column=2, value=sale.date.strftime('%Y-%m-%d'))
        ws.cell(row=row, column=3, value=sale.date.strftime('%H:%M'))
        ws.cell(row=row, column=4, value=sale.customer.name if sale.customer else "Walk-in Customer")
        ws.cell(row=row, column=5, value=sale.get_payment_mode_display())
        ws.cell(row=row, column=6, value=credit_status)
        ws.cell(row=row, column=7, value=items_text)
        ws.cell(row=row, column=8, value=float(sale.total_amount))
        ws.cell(row=row, column=9, value=sale.notes or "")
        
        row += 1
    
    # Adjust column widths
    column_widths = [12, 12, 8, 20, 12, 12, 40, 12, 30]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Add a summary row
    summary_row = row + 2
    ws.cell(row=summary_row, column=1, value="SUMMARY").font = Font(bold=True)
    ws.cell(row=summary_row + 1, column=1, value=f"Total Sales: {sales.count()}")
    ws.cell(row=summary_row + 2, column=1, value=f"Total Amount: ₹{sum(sale.total_amount for sale in sales):.2f}")
    ws.cell(row=summary_row + 3, column=1, value=f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    
    # Create HTTP response with Excel file
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    # Set filename with current date
    filename = f"sales_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Save workbook to response
    wb.save(response)
    
    return response
